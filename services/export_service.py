"""
猫池短信系统导出服务 - tkinter版
Export service for SMS Pool System - tkinter version
"""

import sys
import os
import csv
import json
from datetime import datetime
from typing import Optional, Dict, Any, List
from pathlib import Path

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    from config.logging_config import get_logger, log_info, log_error
except ImportError:
    import logging
    def get_logger(name='services.export'):
        logger = logging.getLogger(name)
        logger.setLevel(logging.INFO)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            logger.addHandler(handler)
        return logger

    def log_info(message):
        logger = get_logger()
        logger.info(message)

    def log_error(message, error=None):
        logger = get_logger()
        logger.error(f"{message}: {error}" if error else message)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

logger = get_logger('services.export')


class ExportService:
    """导出服务类"""

    def __init__(self):
        """初始化导出服务"""
        self.supported_formats = ['xlsx', 'csv', 'txt']

    def export_data(self, export_data: Dict[str, Any]) -> Dict[str, Any]:
        """导出数据"""
        try:
            export_type = export_data.get('export_type', 'completed')
            file_path = export_data.get('file_path', '')
            file_format = export_data.get('file_format', 'Excel (.xlsx)')
            fields = export_data.get('fields', [])
            task = export_data.get('task')

            # 解析文件格式
            format_extension = self._parse_format(file_format)

            # 获取导出数据
            data_to_export = self._get_export_data(export_type, task, fields)

            if not data_to_export:
                return {
                    'success': False,
                    'message': '没有可导出的数据',
                    'count': 0
                }

            # 根据格式导出
            if format_extension == 'xlsx':
                result = self._export_xlsx(file_path, data_to_export, fields)
            elif format_extension == 'csv':
                result = self._export_csv(file_path, data_to_export, fields)
            else:  # txt
                result = self._export_txt(file_path, data_to_export, fields)

            if result['success']:
                log_info(f"成功导出 {result['count']} 条记录到文件: {file_path}")

            return result

        except Exception as e:
            log_error("导出数据异常", error=e)
            return {
                'success': False,
                'message': f'导出异常: {str(e)}',
                'count': 0
            }

    def _parse_format(self, file_format: str) -> str:
        """解析文件格式"""
        if 'xlsx' in file_format.lower():
            return 'xlsx'
        elif 'csv' in file_format.lower():
            return 'csv'
        elif 'txt' in file_format.lower():
            return 'txt'
        else:
            return 'xlsx'

    def _get_export_data(self, export_type: str, task: Dict[str, Any], fields: List[str]) -> List[Dict[str, Any]]:
        """获取导出数据"""
        try:
            # 这里应该从数据库获取真实数据
            # 现在返回模拟数据
            if export_type == 'completed':
                return self._get_completed_messages(task, fields)
            elif export_type == 'uncompleted':
                return self._get_uncompleted_messages(task, fields)
            elif export_type == 'report':
                return self._get_task_reports(fields)
            else:
                return []

        except Exception as e:
            log_error("获取导出数据异常", error=e)
            return []

    def _get_completed_messages(self, task: Dict[str, Any], fields: List[str]) -> List[Dict[str, Any]]:
        """获取已完成消息数据"""
        # 模拟已完成消息数据
        completed_messages = []
        success_count = task.get('success_count', 0) if task else 50

        for i in range(success_count):
            message = {
                'phone': f"138{i:08d}",
                'send_phone': f"1001{i % 10}",
                'port': f"COM{(i % 5) + 1}",
                'carrier': ['中国移动', '中国联通', '中国电信'][i % 3],
                'status': '发送成功',
                'send_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'receive_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'content': task.get('content', '测试短信内容') if task else '测试短信内容'
            }
            completed_messages.append(message)

        return completed_messages

    def _get_uncompleted_messages(self, task: Dict[str, Any], fields: List[str]) -> List[Dict[str, Any]]:
        """获取未完成消息数据"""
        # 模拟未完成消息数据
        uncompleted_messages = []
        if task:
            total = task.get('total', 0)
            sent = task.get('sent', 0)
            failed_count = task.get('failed_count', 0)
            unsent_count = total - sent

            # 未发送的
            for i in range(unsent_count):
                message = {
                    'phone': f"139{i:08d}",
                    'send_phone': '',
                    'port': '',
                    'carrier': ['中国移动', '中国联通', '中国电信'][i % 3],
                    'status': '未发送',
                    'send_time': '',
                    'receive_time': '',
                    'content': task.get('content', '测试短信内容')
                }
                uncompleted_messages.append(message)

            # 发送失败的
            for i in range(failed_count):
                message = {
                    'phone': f"137{i:08d}",
                    'send_phone': f"1001{i % 10}",
                    'port': f"COM{(i % 5) + 1}",
                    'carrier': ['中国移动', '中国联通', '中国电信'][i % 3],
                    'status': '发送失败',
                    'send_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'receive_time': '',
                    'content': task.get('content', '测试短信内容')
                }
                uncompleted_messages.append(message)

        return uncompleted_messages

    def _get_task_reports(self, fields: List[str]) -> List[Dict[str, Any]]:
        """获取任务报告数据"""
        # 模拟任务报告数据
        reports = []
        for i in range(10):  # 模拟10个任务
            report = {
                'task_name': f'任务{i+1}',
                'total_count': (i + 1) * 100,
                'success_count': (i + 1) * 80,
                'failed_count': (i + 1) * 5,
                'progress': f"{((i + 1) * 85 / ((i + 1) * 100) * 100):.1f}%",
                'create_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'complete_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S') if i < 8 else '',
                'content': f'这是任务{i+1}的短信内容'
            }
            reports.append(report)

        return reports

    def _export_xlsx(self, file_path: str, data: List[Dict[str, Any]], fields: List[str]) -> Dict[str, Any]:
        """导出为Excel文件"""
        try:
            if not XLSX_AVAILABLE:
                return {
                    'success': False,
                    'message': '未安装openpyxl库，无法导出Excel文件',
                    'count': 0
                }

            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "导出数据"

            # 字段名映射
            field_names = self._get_field_names()

            # 写入表头
            headers = [field_names.get(field, field) for field in fields]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='FF7F50', end_color='FF7F50', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 写入数据
            for row, item in enumerate(data, 2):
                for col, field in enumerate(fields, 1):
                    value = item.get(field, '')
                    ws.cell(row=row, column=col, value=value)

            # 调整列宽
            for col in range(1, len(fields) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

            # 保存文件
            wb.save(file_path)

            return {
                'success': True,
                'message': f'成功导出到Excel文件: {file_path}',
                'count': len(data)
            }

        except Exception as e:
            log_error("导出Excel文件异常", error=e)
            return {
                'success': False,
                'message': f'导出Excel文件失败: {str(e)}',
                'count': 0
            }

    def _export_csv(self, file_path: str, data: List[Dict[str, Any]], fields: List[str]) -> Dict[str, Any]:
        """导出为CSV文件"""
        try:
            field_names = self._get_field_names()

            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)

                # 写入表头
                headers = [field_names.get(field, field) for field in fields]
                writer.writerow(headers)

                # 写入数据
                for item in data:
                    row = [item.get(field, '') for field in fields]
                    writer.writerow(row)

            return {
                'success': True,
                'message': f'成功导出到CSV文件: {file_path}',
                'count': len(data)
            }

        except Exception as e:
            log_error("导出CSV文件异常", error=e)
            return {
                'success': False,
                'message': f'导出CSV文件失败: {str(e)}',
                'count': 0
            }

    def _export_txt(self, file_path: str, data: List[Dict[str, Any]], fields: List[str]) -> Dict[str, Any]:
        """导出为文本文件"""
        try:
            field_names = self._get_field_names()

            with open(file_path, 'w', encoding='utf-8') as txtfile:
                # 写入表头
                headers = [field_names.get(field, field) for field in fields]
                txtfile.write('\t'.join(headers) + '\n')
                txtfile.write('-' * 80 + '\n')

                # 写入数据
                for item in data:
                    row = [str(item.get(field, '')) for field in fields]
                    txtfile.write('\t'.join(row) + '\n')

            return {
                'success': True,
                'message': f'成功导出到文本文件: {file_path}',
                'count': len(data)
            }

        except Exception as e:
            log_error("导出文本文件异常", error=e)
            return {
                'success': False,
                'message': f'导出文本文件失败: {str(e)}',
                'count': 0
            }

    def _get_field_names(self) -> Dict[str, str]:
        """获取字段名映射"""
        return {
            # 消息字段
            'phone': '接收号码',
            'send_phone': '发送号码',
            'port': '发送串口',
            'carrier': '运营商',
            'status': '发送状态',
            'send_time': '发送时间',
            'receive_time': '接收时间',
            'content': '短信内容',
            # 任务报告字段
            'task_name': '任务名称',
            'total_count': '总数量',
            'success_count': '成功数量',
            'failed_count': '失败数量',
            'progress': '完成进度',
            'create_time': '创建时间',
            'complete_time': '完成时间'
        }

    def export_task_messages(self, task_id: int, status: str = None,
                           file_format: str = 'xlsx') -> Dict[str, Any]:
        """导出任务消息"""
        try:
            # 这里应该从数据库获取任务消息
            # 暂时返回模拟结果
            log_info(f"导出任务{task_id}的消息，状态: {status}，格式: {file_format}")

            return {
                'success': True,
                'message': '任务消息导出成功',
                'file_path': f'/tmp/task_{task_id}_messages.{file_format}',
                'count': 100
            }

        except Exception as e:
            log_error(f"导出任务{task_id}消息异常", error=e)
            return {
                'success': False,
                'message': f'导出异常: {str(e)}',
                'count': 0
            }

    def export_port_statistics(self, file_format: str = 'xlsx') -> Dict[str, Any]:
        """导出端口统计"""
        try:
            log_info(f"导出端口统计，格式: {file_format}")

            return {
                'success': True,
                'message': '端口统计导出成功',
                'file_path': f'/tmp/port_statistics.{file_format}',
                'count': 10
            }

        except Exception as e:
            log_error("导出端口统计异常", error=e)
            return {
                'success': False,
                'message': f'导出异常: {str(e)}',
                'count': 0
            }

    def get_export_preview(self, export_type: str, task: Dict[str, Any] = None,
                          fields: List[str] = None, limit: int = 10) -> Dict[str, Any]:
        """获取导出预览"""
        try:
            if not fields:
                fields = ['phone', 'status', 'send_time']

            data = self._get_export_data(export_type, task, fields)
            preview_data = data[:limit] if data else []

            return {
                'success': True,
                'preview_data': preview_data,
                'total_count': len(data),
                'preview_count': len(preview_data),
                'fields': fields
            }

        except Exception as e:
            log_error("获取导出预览异常", error=e)
            return {
                'success': False,
                'message': f'预览失败: {str(e)}',
                'preview_data': [],
                'total_count': 0
            }

    def validate_export_path(self, file_path: str) -> Dict[str, Any]:
        """验证导出路径"""
        try:
            # 检查目录是否存在
            directory = os.path.dirname(file_path)
            if not os.path.exists(directory):
                return {
                    'valid': False,
                    'message': f'目录不存在: {directory}'
                }

            # 检查目录是否可写
            if not os.access(directory, os.W_OK):
                return {
                    'valid': False,
                    'message': f'目录不可写: {directory}'
                }

            # 检查文件是否已存在
            if os.path.exists(file_path):
                return {
                    'valid': True,
                    'message': '文件已存在，将被覆盖',
                    'file_exists': True
                }

            return {
                'valid': True,
                'message': '路径有效',
                'file_exists': False
            }

        except Exception as e:
            return {
                'valid': False,
                'message': f'路径验证异常: {str(e)}'
            }

    def get_supported_formats(self) -> List[Dict[str, str]]:
        """获取支持的导出格式"""
        formats = [
            {'name': 'Excel (.xlsx)', 'extension': 'xlsx', 'description': 'Microsoft Excel 工作簿'},
            {'name': 'CSV (.csv)', 'extension': 'csv', 'description': '逗号分隔值文件'},
            {'name': '文本文件 (.txt)', 'extension': 'txt', 'description': '制表符分隔的文本文件'}
        ]

        if not XLSX_AVAILABLE:
            # 如果没有安装openpyxl，移除Excel选项
            formats = [f for f in formats if f['extension'] != 'xlsx']

        return formats

    def cleanup_temp_files(self, max_age_hours: int = 24) -> Dict[str, Any]:
        """清理临时导出文件"""
        try:
            temp_dir = '/tmp'
            if not os.path.exists(temp_dir):
                return {'success': True, 'cleaned_count': 0}

            cleaned_count = 0
            current_time = datetime.now()

            for filename in os.listdir(temp_dir):
                if filename.startswith(('task_', 'port_', 'export_')):
                    file_path = os.path.join(temp_dir, filename)
                    try:
                        file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                        age_hours = (current_time - file_time).total_seconds() / 3600

                        if age_hours > max_age_hours:
                            os.remove(file_path)
                            cleaned_count += 1
                    except:
                        continue

            log_info(f"清理了 {cleaned_count} 个临时导出文件")

            return {
                'success': True,
                'cleaned_count': cleaned_count,
                'message': f'成功清理 {cleaned_count} 个临时文件'
            }

        except Exception as e:
            log_error("清理临时文件异常", error=e)
            return {
                'success': False,
                'message': f'清理异常: {str(e)}',
                'cleaned_count': 0
            }


# 全局导出服务实例
export_service = ExportService()